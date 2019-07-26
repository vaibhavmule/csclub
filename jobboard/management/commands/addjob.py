import requests
import pytz
import openpyxl

from datetime import datetime, timedelta

from django.core.management.base import BaseCommand
from django.utils.text import slugify
from jobboard.models import Job, Employer, EmploymentType, AddJobLog


def download_file():
    url = 'https://www.icsi.edu/media/webmodules/Requirement.xlsx'
    local_filename = '/tmp/' + url.split('/')[-1]
    print('Downloading File...', local_filename)
    r = requests.get(url, stream=True)
    with open(local_filename, 'wb') as f:
        for chunk in r.iter_content(chunk_size=1024):
            if chunk:
                f.write(chunk)
    print('Downloaded')
    return local_filename


def sanitize_date(s):
    if isinstance(s, str) and (s.startswith('Reposted') or s.startswith('Re posted on')):
        return s
    if isinstance(s, datetime):
        return s
    if '.' in s:
        return datetime.strptime(s.strip(), '%d.%m.%Y')
    if '/' in s:
        return datetime.strptime(s.strip(), '%d/%m/%Y')
    if '-' in s:
        return datetime.strptime(s.strip(), '%d-%m-%Y')


def max_row(sheet):
    h_column = sheet['G']
    max_col = 2
    now = datetime.now() - timedelta(days=250)
    for col in h_column[2:197]:
        d = sanitize_date(col.value)
        if isinstance(d, str):
            max_col += 1
        elif d.date() >= now.date():
            max_col += 1
        else:
            print(max_col)
            break
    return max_col


def is_latest_file():
    url = 'http://www.icsi.edu/Docs/Webmodules/Requirement.xlsx'
    r = requests.get(url)
    headers = r.headers
    local_tz = pytz.timezone('Asia/Kolkata')
    lst_mod = local_tz.localize(datetime.strptime(
        headers['Last-Modified'],
        '%a, %d %b %Y %X %Z',
    ))
    latest = None
    if AddJobLog.objects.count():
        latest = AddJobLog.objects.latest('created')

    job_log = AddJobLog(
        etag=headers['ETag'].replace('"', ''),
        content_length=int(headers['Content-Length']),
        last_modified=lst_mod)
    job_log.save()
    is_latest = True
    if latest:
        etag = latest.etag == job_log.etag
        content_length = latest.content_length == job_log.content_length
        last_modified = latest.last_modified == job_log.last_modified
        is_downloaded = AddJobLog.objects.filter(
            etag=job_log.etag,
            content_length=job_log.content_length,
            last_modified=job_log.last_modified,
            is_downloaded=True).exists()
        if etag and content_length and last_modified and is_downloaded:
            is_latest = False
    return is_latest, job_log


def add_cs_trainee_job(row):
    """
    S.No.
    Company Name
    Address
    Email id
    Requirement
    Location
    Posted on
    """
    d = row[6].value
    str_d = sanitize_date(row[6].value)
    print(isinstance(str_d, datetime))
    if isinstance(d, datetime):
        print('----------------------------------------\n')
        print("{} Looking for CS Trainees in {}".format(
            row[1].value, row[5].value),)
        html = "<strong>Requirement:</strong> {} <br/><strong>Address:</strong> {}".format(
            row[4].value,
            row[2].value)
        print(html)
        date_posted = datetime.combine(
            d, datetime.now().time())
        employer = None
        slug = slugify(row[1].value)
        if Employer.objects.filter(slug=slug).exists():
            employer = Employer.objects.get(slug=slug)
        else:
            print('here you go')
            employer = Employer.objects.create(title=row[1].value)
        job = Job(
            title="CS Trainee",
            description=html,
            employer=employer,
            location=row[5].value,
            date_posted=date_posted,
            apply_email=row[3].value)
        employment_type, created = EmploymentType.objects.get_or_create(
            title='Internship',
            value='INTERN',
        )
        job.save()
        job.employment_type.add(employment_type)
    elif isinstance(str_d, datetime):
        print('----------------------------------------\n')
        print("{} Looking for CS Trainees in {}".format(
            row[1].value, row[5].value),)
        html = "<strong>Requirement:</strong> {} <br/><strong>Address:</strong> {}".format(
            row[4].value,
            row[2].value)
        print(html)
        date_posted = str_d
        employer = None
        slug = slugify(row[1].value)
        if Employer.objects.filter(slug=slug).exists():
            employer = Employer.objects.get(slug=slug)
        else:
            employer = Employer.objects.create(title=row[1].value)
        job = Job(
            title="CS Trainee",
            description=html,
            employer=employer,
            location=row[5].value,
            date_posted=date_posted,
            apply_email=row[3].value)
        employment_type, created = EmploymentType.objects.get_or_create(
            title='Internship',
            value='INTERN',
        )
        job.save()
        job.employment_type.add(employment_type)
    else:
        print('not-happening---------')


class Command(BaseCommand):
    help = 'Add jobs from ICSI'

    def handle(self, *args, **options):
        # is_latest, job_log = is_latest_file()
        if True:
            file_path = download_file()
            wb = openpyxl.load_workbook(file_path)
            sheet = wb.active
            mx_row = max_row(sheet)
            iter_rows = sheet.iter_rows(min_row=3, max_col=8, max_row=mx_row)
            if mx_row > 2:
                for row in iter_rows:
                    print(row)
                    add_cs_trainee_job(tuple(row))
            # job_log.is_downloaded = True
            # job_log.save()
        self.stdout.write(self.style.SUCCESS('ok'))
